import random
import time

from bs4 import BeautifulSoup
# import nltk
# nltk.download('punkt')
import requests
from twitter import OAuth, Twitter

from openpyxl import load_workbook
import hashlib

# import tokens

from os import environ

tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')

oauth = OAuth(
    environ['ACCESS_KEY'],
    environ['ACCESS_SECRET'],
    environ['CONSUMER_KEY'],
    environ['CONSUMER_SECRET']
)

t = Twitter(auth=oauth)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.87 Safari/537.36',
}

filepath = 'titles.xlsx'

def check_dup(key):
    wb = load_workbook(filepath)
    sheet = wb.active
    max_row = sheet.max_row
    for r in range (1, max_row+1):
        cell_obj = sheet.cell(row = r, column = 1)
        if cell_obj.value == key:
            return True
    return False

def add_key(key):
    wb = load_workbook(filepath)
    sheet = wb.active
    max_row = sheet.max_row
    sheet.cell(row=max_row+1, column=1).value = key
    wb.save(filepath)



def scrape_sd():
    url = 'https://www.sciencedaily.com/news/space_time'
    response = requests.get(url, headers=HEADERS)
    results = BeautifulSoup(response.text, 'html.parser')

    top = results.find(id="heroes")
    
    headlines = top.find_all('a')

    for headline in headlines:
        title = headline.text.strip()
        link = 'https://www.sciencedaily.com' + headline['href']
        tweet = (title, link)
        yield tweet
        #key = hashlib.md5(title.encode())
        #hex_key = key.hexdigest()
        #if (check_dup(hex_key)):
        #    print(check_dup(hex_key))
        #    continue
        #else:
        #    add_key(hex_key)
        #    yield '"%s" %s' % (title, link)
        

def scrape_sn():
    url = 'https://www.sciencenews.org/topic/space'
    response = requests.get(url, headers=HEADERS)
    results = BeautifulSoup(response.text, 'html.parser')

    articles = results.find_all('article')

    for article in articles:
        headline = article.find('h3').find('a')
        title = headline.text.strip()
        link = headline['href']
        tweet = (title, link)
        yield tweet
        
        #key = hashlib.md5(title.encode())
        #hex_key = key.hexdigest()
        #if (check_dup(hex_key)):
        #    continue
        #else:
        #    add_key(hex_key)
        #    yield '"%s" %s' % (title, link)


def tweet():
    print('Start')
    news_sources = ['scrape_sd', 'scrape_sn']
    news_iterators = []
    for source in news_sources:
        news_iterators.append(globals()[source]())

    for i, iterator in enumerate(news_iterators):
        while True:
            try:
                tweet = next(iterator)
                title = tweet[0]
                link = tweet[1]
                key = hashlib.md5(title.encode())
                hex_key = key.hexdigest()
                if (check_dup(hex_key)):
                    #print(check_dup(hex_key))
                    continue
                else:
                    add_key(hex_key)
                    final = '"%s" %s' % (title, link)
                    t.statuses.update(status=final)
                    print(final, end='\n\n')

                time.sleep(600)

            except StopIteration:
                break
                # news_iterators[i] = globals()[news_sources[i]]()

# wb = load_workbook(filepath)
# sheet = wb.active
# max_row = sheet.max_row
# print(max_row)
# if __name__ == "__main__":
#     main()